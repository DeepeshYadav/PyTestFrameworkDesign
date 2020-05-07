import openpyxl


filepath = 'testdata.xlsx'

wb = openpyxl.load_workbook(filepath)

sh = wb.active

data = sh.cell(1, 2)

#print(data.value)

for i in range(2, 5):
    print(sh.cell(i, 1).value)
